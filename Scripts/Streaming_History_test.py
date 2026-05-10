from session_compiler import session_compiler
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.styles import Border, Side, Font, PatternFill

while True:
    user_id = input("First four letters of UUID: ")
    try:
        df = pd.read_csv(f"../Raw Data/Streaming_History_{user_id}.csv").copy()
    except FileNotFoundError:
        print(f"No file found for UUID: {user_id}")
        continue
    break

df = df.rename(columns={
    "master_metadata_track_name": "Track Name",
    "master_metadata_album_artist_name": "Artist Name"
})

df["ts"] = pd.to_datetime(df["ts"])
df["date"] = df["ts"].dt.date
df["time"] = df["ts"].dt.time

most_listened_songs = (
    df.groupby(["Track Name", "Artist Name"])
    .size()
    .rename("Play Count")
)

most_listened_artists = (
    df.groupby(["Artist Name"])
    .size()
    .rename("Play Count")
)

df_skipped = df[
    (df["reason_end"] == "fwdbtn") |
    (df["reason_end"] == "clickrow")
].copy()

skip_counts_songs = (
    df_skipped.groupby(["Track Name", "Artist Name"])
    .size()
    .rename("Skip Count")
)

skip_counts_artists = (
    df_skipped.groupby(["Artist Name"])
    .size()
    .rename("Skip Count")
)

songs_df = pd.concat([most_listened_songs, skip_counts_songs], axis=1).fillna(0)
songs_df["Skip Rate"] = songs_df["Skip Count"] / songs_df["Play Count"] * 100

most_skipped_songs_df = (
    songs_df[songs_df["Play Count"] >= 50]
    .sort_values("Skip Rate", ascending=False)
    .head(50)
    .reset_index()
)

most_listened_songs_df = (
    songs_df.sort_values("Play Count", ascending=False)
    .reset_index()
)

artists_df = pd.concat([most_listened_artists, skip_counts_artists], axis=1).fillna(0)
artists_df["Skip Rate"] = artists_df["Skip Count"] / artists_df["Play Count"] * 100

most_skipped_artists_df = (
    artists_df[artists_df["Play Count"] >= 50]
    .sort_values("Skip Rate", ascending=False)
    .reset_index()
)

most_listened_artists_df = (
    artists_df.sort_values("Play Count", ascending=False)
    .reset_index()
)

past_year_listens = df[df["ts"] >= pd.Timestamp("2025-05-01", tz="UTC")].copy()
past_year_listens["hour"] = past_year_listens["ts"].dt.hour

hourly_plays = (
    past_year_listens.groupby("hour")
    .agg(
        total_plays=("ts", "count"),
        skips=("reason_end", lambda x: x.isin(["fwdbtn", "clickrow"]).sum())
    )
    .reset_index()
    .rename(columns={"hour": "Hour"})
)

hourly_plays["Percentage of Plays"] = hourly_plays["total_plays"] / len(past_year_listens) * 100
hourly_plays["Skip Rate"] = hourly_plays["skips"] / hourly_plays["total_plays"] * 100
hourly_plays = hourly_plays.drop(columns=["total_plays", "skips"])

discovered_songs = df.drop_duplicates(
    subset=["Track Name", "Artist Name"],
    keep="first"
).copy()

discovered_songs["Month"] = discovered_songs["ts"].dt.to_period("M").astype(str)

daily_discovered_songs = (
    discovered_songs.groupby("Month")
    .size()
    .rename("Amount of new songs found")
    .reset_index()
)

def format_pct_column(worksheet, df_out, col_name):
    for cell in worksheet[1]:
        if cell.value == col_name:
            for row in range(2, len(df_out) + 2):
                worksheet.cell(row=row, column=cell.column).number_format = '0.00"%"'
            break

output_path = f"../Song Statistics/{user_id} Song Stats.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    most_skipped_songs_df.to_excel(writer, sheet_name="Most Skipped Songs", index=False)
    most_listened_songs_df.to_excel(writer, sheet_name="Most Played Songs", index=False)
    most_skipped_artists_df.to_excel(writer, sheet_name="Most Skipped Artists", index=False)
    most_listened_artists_df.to_excel(writer, sheet_name="Most Played Artists", index=False)
    daily_discovered_songs.to_excel(writer, sheet_name="No. songs discovered by Month", index=False)
    hourly_plays.to_excel(writer, sheet_name="Plays throughout the day", index=False)

    sheets = {
        "Most Skipped Songs": most_skipped_songs_df,
        "Most Played Songs": most_listened_songs_df,
        "Most Skipped Artists": most_skipped_artists_df,
        "Most Played Artists": most_listened_artists_df,
        "No. songs discovered by Month": daily_discovered_songs,
        "Plays throughout the day": hourly_plays
    }

    for sheet_name, df_out in sheets.items():
        worksheet = writer.sheets[sheet_name]

        format_pct_column(worksheet, df_out, "Skip Rate")
        format_pct_column(worksheet, df_out, "Percentage of Plays")

        for column_cells in worksheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    hourly_worksheet = writer.sheets["Plays throughout the day"]
    hourly_chart = BarChart()
    hourly_chart.title = "Percentage of plays throughout the day"
    hourly_chart.y_axis.title = "Plays"
    hourly_chart.x_axis.title = "Hour of Day"
    hourly_chart.width = 40
    hourly_chart.height = 20
    hourly_chart.legend = None
    hourly_chart.x_axis.number_format = "0"
    hourly_chart.x_axis.tickLblPos = "low"
    hourly_chart.y_axis.number_format = "0"
    hourly_chart.y_axis.tickLblPos = "nextTo"

    h_data = Reference(hourly_worksheet, min_col=2, max_col=2, min_row=1, max_row=25)
    h_labels = Reference(hourly_worksheet, min_col=1, max_col=1, min_row=2, max_row=25)

    hourly_chart.add_data(h_data, titles_from_data=True, from_rows=False)
    hourly_chart.set_categories(h_labels)
    hourly_chart.x_axis.tickLblSkip = 6
    hourly_chart.y_axis.scaling.min = 0
    hourly_chart.y_axis.majorGridlines = None
    hourly_chart.x_axis.delete = False
    hourly_chart.y_axis.delete = False
    hourly_chart.style = 2
    hourly_chart.series[0].graphicalProperties.line.solidFill = "4472C4"
    hourly_chart.series[0].graphicalProperties.line.width = 25000
    hourly_worksheet.add_chart(hourly_chart, "E3")

    worksheet = writer.sheets["No. songs discovered by Month"]
    worksheet.sheet_state = "hidden"

    num_rows = len(daily_discovered_songs) + 1

    chart = LineChart()
    chart.title = "Songs Discovered by Month"
    chart.y_axis.title = "Songs"
    chart.x_axis.title = "Month"
    chart.width = 40
    chart.height = 20
    chart.legend = None
    chart.varyColors = False
    chart.x_axis.number_format = "yyyy-mm"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.numFmt = "0"
    chart.y_axis.tickLblPos = "nextTo"

    data = Reference(worksheet, min_col=2, max_col=2, min_row=1, max_row=num_rows)
    labels = Reference(worksheet, min_col=1, max_col=1, min_row=2, max_row=num_rows)

    chart.add_data(data, titles_from_data=True, from_rows=False)
    chart.set_categories(labels)
    chart.x_axis.tickLblSkip = 6
    chart.y_axis.scaling.min = 0
    chart.y_axis.majorGridlines = None
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.style = 2
    chart.series[0].graphicalProperties.line.solidFill = "4472C4"
    chart.series[0].graphicalProperties.line.width = 25000

    chart_ws = writer.book.create_sheet("Songs Discovered by Month")
    chart_ws.add_chart(chart, "A1")

    thin = Side(style="thin")
    thick = Side(style="thick")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for sheet in writer.book.worksheets:
        if sheet.sheet_state == "hidden":
            continue

        min_row = sheet.min_row
        max_row = sheet.max_row
        min_col = sheet.min_column
        max_col = sheet.max_column

        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value is not None:
                    left = thick if cell.column == min_col else thin
                    right = thick if cell.column == max_col else thin
                    top = thick if cell.row == min_row else thin
                    bottom = thick if cell.row == max_row or cell.row == min_row else thin

                    cell.border = Border(left=left, right=right, top=top, bottom=bottom)

                    if cell.row == min_row:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = header_fill

print("Stats are published")