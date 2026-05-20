from session_compiler import session_compiler
import pandas as pd
import numpy as np
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.chart.legend import Legend
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import timedelta
from openpyxl.styles import PatternFill


# ── Analysis function ─────────────────────────────────────────────────────────

def analyse(f_df, a_df, skip_no):

    most_listened_songs = (
        f_df.groupby(["Track Name", "Artist Name"])
        .size()
        .rename("Play Count")
    )

    most_listened_artists = (
        f_df.groupby(["Artist Name"])
        .size()
        .rename("Play Count")
    )

    df_skipped = f_df[
        (f_df["reason_end"] == "fwdbtn") |
        (f_df["reason_end"] == "clickrow")
    ].copy()

    skip_counts_songs = (
        df_skipped.groupby(["Track Name", "Artist Name"])
        .size()
        .rename("Skip Count")
    )

    skip_counts_artists = (
        df_skipped.groupby(["Artist Name"])
        .size()
        .rename("Skip Count")
    )

    songs_df = pd.concat([most_listened_songs, skip_counts_songs], axis=1).fillna(0)
    songs_df["Skip Rate"] = songs_df["Skip Count"] / songs_df["Play Count"] * 100

    most_skipped_songs_df = (
        songs_df[songs_df["Play Count"] >= skip_no]
        .sort_values(["Skip Rate","Play Count"], ascending=[False,False])
        .reset_index()
    )

    least_skipped_songs_df = (
        songs_df[songs_df["Play Count"] >= skip_no]
        .sort_values(["Skip Rate","Play Count"], ascending=[True,False])
        .reset_index()
    )

    most_listened_songs_df = (
        songs_df.sort_values("Play Count", ascending=False)
        .reset_index()
    )

    artists_df = pd.concat([most_listened_artists, skip_counts_artists], axis=1).fillna(0)
    artists_df["Skip Rate"] = artists_df["Skip Count"] / artists_df["Play Count"] * 100

    most_skipped_artists_df = (
        artists_df[artists_df["Play Count"] >= skip_no]
        .sort_values(["Skip Rate","Play Count"], ascending=[False,False])
        .reset_index()
    )

    least_skipped_artists_df = (
        artists_df[artists_df["Play Count"] >= skip_no]
        .sort_values(["Skip Rate","Play Count"], ascending=[True,False])
        .reset_index()
    )

    most_listened_artists_df = (
        artists_df.sort_values("Play Count", ascending=False)
        .reset_index()
    )

    audio_features = ["danceability", "energy", "speechiness",
                      "acousticness", "instrumentalness", "liveness", "valence", "loudness",
                      "tempo", "key", "mode","popularity"]

    df_m = a_df.copy()
    df_m["Month"] = df_m["ts"].dt.to_period("M").astype(str)

    a_songs = a_df.groupby(["Track Name", "Artist Name"])["popularity"].mean()
    
    a_songs = a_df.groupby(["Track Name", "Artist Name"])["popularity"].mean()
    song_plays = f_df.groupby(["Track Name", "Artist Name"]).size().rename("Play Count")

    a_songs_df = pd.concat([a_songs, song_plays], axis=1).dropna()
    a_songs_df = a_songs_df[a_songs_df["Play Count"] >= skip_no/2]

    most_popular_song = a_songs_df.sort_values("popularity", ascending=False).head(1)
    most_popular_song = f"{most_popular_song.index[0][0]}, {most_popular_song.index[0][1]}"

    least_popular_song = a_songs_df.sort_values("popularity", ascending=True).head(1)
    least_popular_song = f"{least_popular_song.index[0][0]}, {least_popular_song.index[0][1]}"
    
    
    a_artists = a_df.groupby("Artist Name")["popularity"].mean()
    artist_plays = f_df.groupby("Artist Name").size().rename("Play Count")

    a_artists_df = pd.concat([a_artists, artist_plays], axis=1).dropna()
    a_artists_df = a_artists_df[a_artists_df["Play Count"] >= skip_no/2]

    most_popular_artist = a_artists_df.sort_values("popularity", ascending=False).head(1)
    most_popular_artist = most_popular_artist.index[0]
    least_popular_artist = a_artists_df.sort_values("popularity", ascending=True).head(1)
    least_popular_artist = least_popular_artist.index[0]

    monthly_moods = (
        df_m.groupby("Month")
        .agg(**{feature: (feature, "mean") for feature in audio_features})
        .reset_index()
    )

    monthly_moods = monthly_moods[["Month"] + audio_features]
    mood_features_to_smooth = ["danceability", "energy", "speechiness",
                               "acousticness", "instrumentalness", "liveness", "valence"]

    monthly_moods[mood_features_to_smooth] = (
        monthly_moods[mood_features_to_smooth]
        .rolling(window=12, center=True, min_periods=1)
        .mean()
    )

    discovered_songs = a_df.drop_duplicates(
        subset=["Track Name", "Artist Name"],
        keep="first"
    ).copy()

    discovered_songs["Month"] = discovered_songs["ts"].dt.to_period("M").astype(str)

    daily_discovered_songs = (
        discovered_songs.groupby("Month")
        .size()
        .rename("Amount of new songs found")
        .reset_index()
    )

    Month_analysis = daily_discovered_songs.merge(
        monthly_moods[["Month"] + audio_features],
        on="Month",
        how="left"
    )

    skip_rate = len(df_skipped) / len(f_df)
    Daily_plays = (
        f_df.groupby("date")["ms_played"].sum() / 60000
    ).rename("Minutes Listened").reset_index()

    date_range = pd.date_range(f_df["date"].min(), f_df["date"].max())
    listens_per_day = Daily_plays["Minutes Listened"].sum() / len(date_range)

    if songs_df[(songs_df["Skip Rate"] == 0) & (songs_df["Play Count"] >= skip_no)].empty:
        most_loved_song = songs_df.sort_values("Skip Rate", ascending=True).head(1)
    else:
        most_loved_song = (
            songs_df[(songs_df["Skip Rate"] == 0) & (songs_df["Play Count"] >= skip_no)]
            .sort_values("Play Count", ascending=False)
            .head(1)
        )

    artist_day_counts = f_df.groupby("Artist Name")["date"].nunique().rename("Days Listened")
    artists_df = artists_df.join(artist_day_counts)

    loved_candidates = artists_df[
        (artists_df["Skip Rate"] == 0) &
        (artists_df["Play Count"] >= skip_no) &
        (artists_df["Days Listened"] >= 2)
    ]

    if loved_candidates.empty:
        most_loved_artist = artists_df[artists_df["Play Count"] >= skip_no].sort_values("Skip Rate", ascending=True).head(1)
    else:
        most_loved_artist = loved_candidates.sort_values("Play Count", ascending=False).head(1)

    least_loved_song = songs_df[songs_df["Play Count"] >= skip_no].sort_values("Skip Rate", ascending=False).head(1)
    least_loved_artist = artists_df[artists_df["Play Count"] >= skip_no].sort_values("Skip Rate", ascending=False).head(1)

    if most_loved_artist.empty:
        most_loved_artist_val = "N/A"
    else:
        most_loved_artist_val = most_loved_artist.index[0]

    if least_loved_artist.empty:
        least_loved_artist_val = "N/A"
    else:
        least_loved_artist_val = least_loved_artist.index[0]

    if most_loved_song.empty:
        most_loved_song_val = "N/A"
    else:
        most_loved_song_val = f"{most_loved_song.index[0][0]}, {most_loved_song.index[0][1]}"

    if least_loved_song.empty:
        least_loved_song_val = "N/A"
    else:
        least_loved_song_val = f"{least_loved_song.index[0][0]}, {least_loved_song.index[0][1]}"

    # Longest session
    f_sorted = f_df.sort_values("ts").reset_index(drop=True)
    streaming_sessions = {}
    session_no = 1
    longest_session = []

    for j in range(len(f_sorted) - 1):
        streaming_sessions.setdefault(session_no, []).append(f_sorted.iloc[j].to_dict())
        if f_sorted.iloc[j + 1]["ts"] - f_sorted.iloc[j]["ts"] >= timedelta(minutes=30):
            current = pd.DataFrame(streaming_sessions[session_no])
            if current["ms_played"].sum() > (pd.DataFrame(longest_session)["ms_played"].sum() if longest_session else 0):
                longest_session = streaming_sessions[session_no]
            session_no += 1

    streaming_sessions.setdefault(session_no, []).append(f_sorted.iloc[-1].to_dict())
    current = pd.DataFrame(streaming_sessions[session_no])
    if current["ms_played"].sum() > (pd.DataFrame(longest_session)["ms_played"].sum() if longest_session else 0):
        longest_session = streaming_sessions[session_no]

    longest_session_df = pd.DataFrame(longest_session)
    longest_session_date = longest_session_df["date"].iloc[0]
    longest_session_hours = round(longest_session_df["ms_played"].sum() / 60000 / 60, 1)

    summary_df = pd.DataFrame({
        "Statistics": [
            "Average Minutes Listened Per Day",
            "Skip Rate",
            "Longest Listening Session",
            "Number 1 artist",
            "Number 1 song",
            "My no skip artist",
            "Artists you Skip but can't Quit",
            "Full plays only",
            "Just can't get through this song",
            "Mainstream Moment",
            "Underground Gem",
            "Biggest Name in your Library",
            "Best Kept Secret"
        ],
        "Results": [
            f"{int(listens_per_day)} minutes",
            f"{round(skip_rate * 100, 2)}%",
            f"{longest_session_date}, {round(longest_session_hours,1)} hours",
            most_listened_artists_df.iloc[0]["Artist Name"],
            f"{most_listened_songs_df.iloc[0]['Track Name']}, {most_listened_songs_df.iloc[0]['Artist Name']}",
            most_loved_artist_val,
            least_loved_artist_val,
            most_loved_song_val,
            least_loved_song_val,
            most_popular_song,
            least_popular_song,
            most_popular_artist,
            least_popular_artist
        ]
    })

    return {
        "Listening Overview": summary_df,
        "Most Skipped Songs": most_skipped_songs_df,
        "Least Skipped Songs": least_skipped_songs_df,
        "Most Played Songs": most_listened_songs_df,
        "Most Skipped Artists": most_skipped_artists_df,
        "Least Skipped Artists": least_skipped_artists_df,
        "Most Played Artists": most_listened_artists_df,
        "Plays by Month": Month_analysis,
    }


# ── Side by side writer ───────────────────────────────────────────────────────

def write_side_by_side(worksheet, df_left, df_right, gap=2):
    left_cols = len(df_left.columns)
    right_start_col = left_cols + gap + 1
    right_cols = len(df_right.columns)

    # Merge and write headers
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=left_cols)
    worksheet.cell(row=1, column=1, value="Previous Year")

    worksheet.merge_cells(start_row=1, start_column=right_start_col, end_row=1, end_column=right_start_col + right_cols - 1)
    worksheet.cell(row=1, column=right_start_col, value="All Time")

    # Write tables starting from row 2
    rows_left = list(dataframe_to_rows(df_left, index=False, header=True))
    for r_idx, row in enumerate(rows_left, 2):
        for c_idx, value in enumerate(row, 1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)

    col_offset = len(df_left.columns) + gap
    rows_right = list(dataframe_to_rows(df_right, index=False, header=True))
    for r_idx, row in enumerate(rows_right, 2):
        for c_idx, value in enumerate(row, 1):
            worksheet.cell(row=r_idx, column=c_idx + col_offset, value=value)


# ── Formatting ────────────────────────────────────────────────────────────────

def format_sheet(worksheet, df_left, df_right, gap=2):
    thin = Side(style="thin")
    thick = Side(style="thick")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_fill = PatternFill(start_color="2B4A8A", end_color="2B4A8A", fill_type="solid")

    left_cols = len(df_left.columns)
    right_start_col = left_cols + gap + 1
    right_cols = len(df_right.columns)

    # Style title row
    for col in [1, right_start_col]:
        cell = worksheet.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center")

    for c in range(1, left_cols + 1):
        cell = worksheet.cell(row=1, column=c)
        left = thick if c == 1 else Side(style=None)
        right = thick if c == left_cols else Side(style=None)
        cell.border = Border(left=left, right=right, top=thick, bottom=thick)

    # Border for right title merged cell
    for c in range(right_start_col, right_start_col + right_cols):
        cell = worksheet.cell(row=1, column=c)
        left = thick if c == right_start_col else Side(style=None)
        right = thick if c == right_start_col + right_cols - 1 else Side(style=None)
        cell.border = Border(left=left, right=right, top=thick, bottom=thick)

    def format_table(min_col, max_col, num_rows):
        for r in range(2, num_rows + 3):
            for c in range(min_col, max_col + 1):
                cell = worksheet.cell(row=r, column=c)
                cell.alignment = Alignment(horizontal="center")
                if cell.value is not None:
                    left = thick if c == min_col else thin
                    right = thick if c == max_col else thin
                    top = thick if r == 2 else thin
                    bottom = thick if r == num_rows + 2 or r == 2 else thin
                    cell.border = Border(left=left, right=right, top=top, bottom=bottom)
                    if r == 2:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = header_fill

    num_rows = max(len(df_left), len(df_right))
    format_table(1, left_cols, num_rows)
    format_table(right_start_col, right_start_col + right_cols - 1, num_rows)

    from openpyxl.utils import get_column_letter

    max_col = worksheet.max_column
    for c_idx in range(1, max_col + 1):
        max_length = 0
        for row in worksheet.iter_rows(min_col=c_idx, max_col=c_idx):
            for cell in row:
                if cell.value is not None and not isinstance(cell, type(worksheet.cell(1,1).__class__)):
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
        worksheet.column_dimensions[get_column_letter(c_idx)].width = min(max_length + 2, 50)


def format_pct_column(worksheet, col_name, df, col_offset=0):
    for cell in worksheet[2]:
        if cell.value == col_name:
            for row in range(3, len(df) + 3):
                worksheet.cell(row=row, column=cell.column + col_offset).number_format = '0.00"%"'
            break


# ── Main ──────────────────────────────────────────────────────────────────────

def main(user_id):
    while True:
        try:
            a_df = pd.read_csv(f"../Volunteer Data/RAW/Audio_Streaming_History_{user_id}.csv").copy()
            f_df = pd.read_csv(f"../Volunteer Data/RAW/Streaming_history_{user_id}.csv").copy()
        except FileNotFoundError:
            print(f"No file found for name: {user_id}")
            continuex
        break

    f_df = f_df.rename(columns={
        "master_metadata_track_name": "Track Name",
        "master_metadata_album_artist_name": "Artist Name"
    })
    a_df = a_df.rename(columns={
        "master_metadata_track_name": "Track Name",
        "master_metadata_album_artist_name": "Artist Name"
    })

    song_count_a =len(a_df[["Track Name","Artist Name"]].drop_duplicates())
    song_count_f = len(f_df[["Track Name","Artist Name"]].drop_duplicates())
    lost_songs = song_count_f - song_count_a
    print(f'Number of songs lost due to AF mapping: {lost_songs} out of {song_count_f},({round((lost_songs/song_count_f)*100,2)}%)')
    f_df["Artist Name"] = f_df["Artist Name"].replace("Electric Light Orchestra", "ELO")

    a_df["ts"] = pd.to_datetime(a_df["ts"])
    a_df["date"] = a_df["ts"].dt.date
    a_df["time"] = a_df["ts"].dt.time

    f_df["ts"] = pd.to_datetime(f_df["ts"])
    f_df["date"] = f_df["ts"].dt.date
    f_df["time"] = f_df["ts"].dt.time

    # Run analysis for both periods
    all_time = analyse(f_df, a_df, 30)

    prev_year_cutoff = pd.Timestamp("2025-04-01", tz="UTC")
    prev_year = analyse(
        f_df[f_df["ts"] >= prev_year_cutoff].copy(),
        a_df[a_df["ts"] >= prev_year_cutoff].copy(),
        10
    )

    # Hourly plays (all time only — not duplicated)
    past_year_listens = a_df[a_df["ts"] >= prev_year_cutoff].copy()
    past_year_listens["hour"] = past_year_listens["ts"].dt.hour

    audio_features = ["danceability", "energy", "speechiness",
                      "acousticness", "instrumentalness", "liveness", "valence", "loudness",
                      "tempo", "key", "mode"]

    hourly_plays = (
        past_year_listens.groupby("hour")
        .agg(
            total_plays=("ts", "count"),
            skips=("reason_end", lambda x: x.isin(["fwdbtn", "clickrow"]).sum()),
            **{feature: (feature, "mean") for feature in audio_features}
        )
        .reset_index()
        .rename(columns={"hour": "Hour"})
    )

    hourly_plays["Percentage of Plays"] = hourly_plays["total_plays"] / len(past_year_listens) * 100
    hourly_plays["Skip Rate"] = hourly_plays["skips"] / hourly_plays["total_plays"] * 100
    hourly_plays = hourly_plays.drop(columns=["total_plays", "skips"])
    hourly_plays = hourly_plays[["Hour", "Percentage of Plays", "Skip Rate"] + audio_features]
    hour_order = list(range(5, 24)) + list(range(0, 5))
    hourly_plays = hourly_plays.set_index("Hour").reindex(hour_order).reset_index()

    output_path = f"../Volunteer Data/Spotify Stats/{user_id}'s Spotify Stats.xlsx"

    sheet_names = [
        "Listening Overview",
        "Most Skipped Songs",
        "Least Skipped Songs",
        "Most Played Songs",
        "Most Skipped Artists",
        "Least Skipped Artists",
        "Most Played Artists",
        "Plays by Month",
    ]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # Write side-by-side sheets
        for sheet_name in sheet_names:
            df_left = prev_year[sheet_name]
            df_right = all_time[sheet_name]

            # Create sheet manually
            writer.book.create_sheet(sheet_name)
            worksheet = writer.sheets[sheet_name]

            # Hide Plays by Month
            if sheet_name == "Plays by Month":
                worksheet.sheet_state = "hidden"

            write_side_by_side(worksheet, df_left, df_right, gap=2)
            format_sheet(worksheet, df_left, df_right, gap=2)

            # Format pct columns for both tables
            format_pct_column(worksheet, "Skip Rate", df_left)
            format_pct_column(worksheet, "Percentage of Plays", df_left)
            format_pct_column(worksheet, "Skip Rate", df_right, col_offset=len(df_left.columns) + 2)
            format_pct_column(worksheet, "Percentage of Plays", df_right, col_offset=len(df_left.columns) + 2)

        # Write hourly plays sheet normally
        hourly_plays.to_excel(writer, sheet_name="Plays throughout the day", index=False)
        hourly_worksheet = writer.sheets["Plays throughout the day"]

        for column_cells in hourly_worksheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            hourly_worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 50)

        # Hourly bar chart
        hourly_chart = BarChart()
        hourly_chart.title = "Percentage of plays throughout the day"
        hourly_chart.y_axis.title = "Plays"
        hourly_chart.x_axis.title = "Hour of Day"
        hourly_chart.width = 40
        hourly_chart.height = 20
        hourly_chart.legend = None
        hourly_chart.x_axis.number_format = "0"
        hourly_chart.x_axis.tickLblPos = "low"
        hourly_chart.y_axis.number_format = "0"
        hourly_chart.y_axis.tickLblPos = "nextTo"

        h_data = Reference(hourly_worksheet, min_col=2, max_col=2, min_row=1, max_row=25)
        h_labels = Reference(hourly_worksheet, min_col=1, max_col=1, min_row=2, max_row=25)
        hourly_chart.add_data(h_data, titles_from_data=True, from_rows=False)
        hourly_chart.set_categories(h_labels)
        hourly_chart.x_axis.tickLblSkip = 6
        hourly_chart.y_axis.scaling.min = 0
        hourly_chart.y_axis.majorGridlines = None
        hourly_chart.x_axis.delete = False
        hourly_chart.y_axis.delete = False
        hourly_chart.style = 2
        hourly_chart.series[0].graphicalProperties.line.solidFill = "4472C4"
        hourly_chart.series[0].graphicalProperties.line.width = 25000
        hourly_worksheet.add_chart(hourly_chart, "A27")

        # Mood line chart
        mood_chart = LineChart()
        mood_chart.title = "Moods throughout the day"
        mood_chart.y_axis.title = "Audio feature score"
        mood_chart.x_axis.title = "Hour of Day"
        mood_chart.width = 40
        mood_chart.height = 20
        mood_chart.legend = Legend()
        mood_chart.x_axis.number_format = "0"
        mood_chart.x_axis.tickLblPos = "low"
        mood_chart.y_axis.number_format = "0.00"
        mood_chart.y_axis.tickLblPos = "nextTo"

        m_data = Reference(hourly_worksheet, min_col=4, max_col=10, min_row=1, max_row=25)
        m_labels = Reference(hourly_worksheet, min_col=1, max_col=1, min_row=2, max_row=25)
        mood_chart.add_data(m_data, titles_from_data=True, from_rows=False)
        mood_chart.set_categories(m_labels)
        mood_chart.x_axis.tickLblSkip = 6
        mood_chart.y_axis.scaling.min = 0
        mood_chart.y_axis.majorGridlines = None
        mood_chart.x_axis.delete = False
        mood_chart.y_axis.delete = False
        mood_chart.style = 2
        mood_chart.series[0].graphicalProperties.line.solidFill = "4472C4"
        mood_chart.series[0].graphicalProperties.line.width = 25000
        hourly_worksheet.add_chart(mood_chart, "A66")
        mood_chart.legend.layout = Layout(
            manualLayout=ManualLayout(x=0.85, y=0, w=0.15, h=0.4, xMode="factor", yMode="factor")
        )
        mood_chart.legend.overlay = False

        # Songs discovered chart (from hidden Plays by Month sheet)
        plays_by_month_ws = writer.sheets["Plays by Month"]
        num_rows = len(all_time["Plays by Month"]) + 1

        chart = LineChart()
        chart.title = "Songs Discovered by Month"
        chart.y_axis.title = "Songs"
        chart.x_axis.title = "Month"
        chart.width = 40
        chart.height = 20
        chart.legend = None
        chart.varyColors = False
        chart.x_axis.number_format = "yyyy-mm"
        chart.x_axis.tickLblPos = "low"
        chart.y_axis.numFmt = "0"
        chart.y_axis.tickLblPos = "nextTo"

        data = Reference(plays_by_month_ws, min_col=18, max_col=18, min_row=2, max_row=num_rows)
        labels = Reference(plays_by_month_ws, min_col=17, max_col=17, min_row=2, max_row=num_rows)
        chart.add_data(data, titles_from_data=True, from_rows=False)
        chart.set_categories(labels)
        chart.x_axis.tickLblSkip = 6
        chart.y_axis.scaling.min = 0
        chart.y_axis.majorGridlines = None
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.style = 2
        chart.series[0].graphicalProperties.line.solidFill = "4472C4"
        chart.series[0].graphicalProperties.line.width = 25000

        chart_ws = writer.book.create_sheet("Songs played by Month")
        chart_ws.add_chart(chart, "A1")

        m_mood_chart = LineChart()
        m_mood_chart.title = "Moods throughout History"
        m_mood_chart.y_axis.title = "Audio feature score"
        m_mood_chart.x_axis.title = "Month"
        m_mood_chart.width = 40
        m_mood_chart.height = 20
        m_mood_chart.legend = Legend()
        m_mood_chart.x_axis.number_format = "0"
        m_mood_chart.x_axis.tickLblPos = "low"
        m_mood_chart.y_axis.number_format = "0.00"
        m_mood_chart.y_axis.tickLblPos = "nextTo"

        m_m_data = Reference(plays_by_month_ws, min_col=19, max_col=25, min_row=2, max_row=num_rows)
        m_m_labels = Reference(plays_by_month_ws, min_col=17, max_col=17, min_row=2, max_row=num_rows)
        m_mood_chart.add_data(m_m_data, titles_from_data=True, from_rows=False)
        m_mood_chart.set_categories(m_m_labels)
        m_mood_chart.x_axis.tickLblSkip = 6
        m_mood_chart.y_axis.scaling.min = 0
        m_mood_chart.y_axis.majorGridlines = None
        m_mood_chart.x_axis.delete = False
        m_mood_chart.y_axis.delete = False
        m_mood_chart.style = 2
        m_mood_chart.series[0].graphicalProperties.line.solidFill = "4472C4"
        m_mood_chart.series[0].graphicalProperties.line.width = 25000
        chart_ws.add_chart(m_mood_chart, "A40")
        m_mood_chart.legend.layout = Layout(
            manualLayout=ManualLayout(x=0.85, y=0, w=0.15, h=0.4, xMode="factor", yMode="factor")
        )
        m_mood_chart.legend.overlay = False

        no_fill = PatternFill(fill_type=None)

        for sheet in writer.book.worksheets:
            sheet.sheet_view.showGridLines = False

    print("Stats are published")


if __name__ == "__main__":
    main()